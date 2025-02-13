from django.shortcuts import render, HttpResponse
from .models import User, ExpenseCategory, ExpenseClaim
from rest_framework import viewsets
from .serializers import UserSerializer, ExpenseClaimSerializer, ExpenseCategorySerializer,ClaimRequestSerializer
from .permissions import *
from rest_framework.response import Response
from rest_framework import status
from rest_framework.exceptions import PermissionDenied

from rest_framework.decorators import action

# for report
import openpyxl
from io import BytesIO
from django.db.models import Sum

# Create your views here.
class UserView(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [ReadOnlyForAnonymous, NoCreateDeleteForEmployeeAndManager, OwnDataUpdateOnly]
    
    def create(self, request, *args, **kwargs):
        a = super().create(request, *args, **kwargs)
        return Response({'msg':'created successfully'}, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        a = super().update(request, *args, **kwargs)
        return Response({'msg':'updated successfully'}, status=status.HTTP_200_OK)
    
    def destroy(self, request, *args, **kwargs):
        a = super().destroy(request, *args, **kwargs)
        return Response({'msg':'deleted successfully'}, status=status.HTTP_204_NO_CONTENT)
        
    
class ExpenseClaimView(viewsets.ModelViewSet):
    serializer_class = ExpenseClaimSerializer
    permission_classes = [NoAccessForAnonymous]
    
    def get_queryset(self):
        queryset_data = ExpenseClaim.objects.all()
        if self.request.user.is_authenticated and (self.request.user.is_employee or self.request.user.is_manager):
            queryset_data = queryset_data.filter(employee=self.request.user)
        return queryset_data
    
    def filter_queryset(self, queryset):
        status = self.request.query_params.get('status')
        expense_type = self.request.query_params.get('expense_type')
        if status:
            queryset = queryset.filter(status=status)
        if expense_type:
            queryset = queryset.filter(expense_type__expense_category=expense_type)
        return queryset
    
    def list(self, request, *args, **kwargs):
        a = super().list(request, args, kwargs)
        queryset = self.get_queryset()
        if queryset:
            return a
        return Response({'msg':'No Any Claims'})
    
    def destroy(self, request, *args, **kwargs):
        try:
            expense = ExpenseClaim.objects.get(pk=kwargs.get('pk'))
        except ExpenseClaim.DoesNotExist:
            return Response({'msg':"You don't have any claim with such id"}, status=status.HTTP_400_BAD_REQUEST)
        if expense.employee == request.user:
            if expense.status != 'pending':
                raise PermissionDenied({'msg':'This claim cannot be updated or deleted because it has been approved or rejected'})
        return super().destroy(request, args, kwargs)
    
        
class ExpenseCategoryView(viewsets.ModelViewSet):
    queryset = ExpenseCategory.objects.all()
    serializer_class = ExpenseCategorySerializer
    permission_classes = [NoAccessForAnonymous, IsaAdminOrReadOnly]
    

class ClaimRequestView(viewsets.ModelViewSet):
    serializer_class = ClaimRequestSerializer
    permission_classes = [NoAccessForAnonymous, NoAccessForEmployee, NoCreateAndDelPerm]
    
    def get_queryset(self):
        if self.request.user.is_admin:
            return ExpenseClaim.objects.all()
        allusers = User.objects.filter(supervisor=self.request.user)
        # print(allusers)
        allclaimreq = ExpenseClaim.objects.filter(employee__in=allusers)
        return allclaimreq
    
    def filter_queryset(self, queryset):
        status = self.request.query_params.get('status')
        expense_type = self.request.query_params.get('expense_type')
        if status:
            queryset = queryset.filter(status=status)
        if expense_type:
            queryset = queryset.filter(expense_type__expense_category=expense_type)
        return queryset
    
    @action(detail=False, methods=['GET'], permission_classes=[AdminOnly])
    def view_report(self, request):
        employee_id = request.query_params.get('employee_id')
        if employee_id is not None:
            try:
                emp = User.objects.get(id=employee_id)
            except User.DoesNotExist:
                return Response({'msg':"No such user with the given id"}, status=status.HTTP_400_BAD_REQUEST)
            allEmpClaims = ExpenseClaim.objects.filter(employee=emp)
        else:
            allEmpClaims = ExpenseClaim.objects.all()
        serialize = ExpenseClaimSerializer(allEmpClaims, many=True, context = {'request': request})
        return Response(serialize.data)
        
    @action(detail=False, methods=['GET'], permission_classes=[AdminOnly])
    def generate_report(self, request):
        employee_id = request.query_params.get('employee_id')
        if employee_id is not None:
            try:
                emp = User.objects.get(id=employee_id)
            except User.DoesNotExist:
                return Response({'msg':"No such user with the given id"}, status=status.HTTP_400_BAD_REQUEST)
            allEmpClaims = ExpenseClaim.objects.filter(employee=emp)
        else:
            allEmpClaims = ExpenseClaim.objects.all()
        serialize = ExpenseClaimSerializer(allEmpClaims, many=True, context = {'request': request})
        serialized_data = serialize.data
        # print(serialized_data)
        
        total_pending_amt = allEmpClaims.filter(status='pending').aggregate(amt=Sum('amount'))
        total_approved_amt = allEmpClaims.filter(status='approved').aggregate(amt=Sum('amount'))
        total_rejected_amt = allEmpClaims.filter(status='rejected').aggregate(amt=Sum('amount'))
        pending = total_pending_amt['amt']
        approved = total_approved_amt['amt']
        rejected = total_rejected_amt['amt']
        
        # ---------- for downloading excel file ----------
        wb = openpyxl.Workbook()
        sheet1 = wb.active
        # for title
        sheet1.merge_cells('A1:D2')
        sheet1['A1'] = f'Users Expense Report'
        sheet1['A1'].font = openpyxl.styles.Font(bold=True)
        sheet1['A1'].alignment = openpyxl.styles.Alignment(horizontal='center',vertical='center')
        sheet1['A1'].fill = openpyxl.styles.PatternFill(start_color="FF6347", end_color="FFF000", fill_type="solid")
        sheet1['A3']='Employee ID'
        sheet1['B3']='Amount'
        sheet1['C3']='Status'
        sheet1['D3']='Expense Type'
        # for records
        count = 0
        for count, data in enumerate(serialized_data):
            row = count+4
            sheet1[f'A{row}']=data['id']
            sheet1[f'B{row}']=data['amount']
            sheet1[f'C{row}']=data['status']
            sheet1[f'D{row}']=data['expense_type']
            count += 1
        pending_cell = count+5
        approved_cell = count+6
        rejected_cell = count+7
        sheet1[f'A{pending_cell}']=f'Total Amount(pending): {pending}'
        sheet1[f'A{approved_cell}']=f'Total Amount(approved): {approved}'
        sheet1[f'A{rejected_cell}']=f'Total Amount(rejected): {rejected}'
        
        # -- Taken from stactoverflow --
        # Write the Excel file to an in-memory buffer
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)  # Go to the beginning of the file in memory
        # Prepare the HTTP response to send the file to the user
        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=excelfile.xlsx'
        return response
        # -- Taken from stactoverflow --
        # ---------- for downloading excel file ----------

        